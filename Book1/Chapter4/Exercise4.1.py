####Playing with different file extensions
import csv

#tsv_source_file = open("data/ShugermanProsecutorPoliticians-SupremeCourtJustices.tsv","r")
#politicians_reader = csv.DictReader(tsv_source_file,delimiter='\t') #By changin "delimiter" attribute we can open other types of table-based data formats
#print(next(politicians_reader)) #We could rename the file to an .txt extension and the output of these lines won't change.



from openpyxl import load_workbook

source_workbook = load_workbook(filename="data/fredgraph.xlsx") #an xlsx file can have multiple sheets, not this case
print(source_workbook.sheetnames)

for sheet_num, sheet_name in enumerate(source_workbook.sheetnames): #Recorremos cada hoja del libro
    current_sheet = source_workbook[sheet_name]
    print(sheet_name) #Imprimimos el nombre de cada una

    output_file = open("xlsx_"+sheet_name+".csv","w") #Now we see a different utility of the built-in function open()
    output_writer = csv.writer(output_file) #Convertimos cada hoja en un archivo csv

    for row in current_sheet.iter_rows(): #Recorremos cada fila de cada hoja, una a una
        row_cells = []
        

        for cell in row: #Para el valor de cada columna en cada fila. Necesitamos hacer esto ya que la librería openpyxl trata cada fila de la hoja como una tupla
            row_cells.append(cell.value)
            output_writer.writerow(row_cells) #Escribe las nuevas filas en el fichero
    
    output_file.close() #Cerramos el archivo


